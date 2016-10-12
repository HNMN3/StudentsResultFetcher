# Keep coding and change the world..And do not forget anything..Not Again..
import pickle

import openpyxl

headings = ['Name', 'COMPUTER NETWORKS', 'DESIGN AND ANALYSIS OF ALGORITHMS', 'THEORY OF COMPUTATION',
            'COMPUTER GRAPHICS & MULTIMEDIA TECHNIQUES', 'EMBEDDED SYSTEM DESIGN',
            'ADVANCE TOPICS IN OPERATING SYSTEMS', 'JAVA PROGRAMMING LAB', 'Computer graphics & multimedia lab',
            'DESIGN AND ANALYSIS OF ALGORITHMS Lab', 'Embedded System Design Lab', 'Humanities and Social Sciences',
            'Discipline', 'Total Marks', 'Result']


def save_to_excel(filename, data):
    wb = openpyxl.Workbook()
    ws = wb.get_sheet_by_name('Sheet')
    i = 1
    j = 65
    for item in headings:
        ws[chr(j) + str(i)] = item
        j += 1
    l = len(headings)
    for row in data:
        i += 1
        j = 65
        if len(row) < l:
            row.insert(-2, 0)
        for column in row:
            try:
                cell_data = (
                    int(column[0].replace('*', '')) + int(column[1].replace('*', '')) if type(
                        column) is list else column)
            except Exception:
                cell_data = 0
            ws[chr(j) + str(i)] = str(cell_data) if cell_data != 'AA' else 0
            j += 1
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 4

    wb.save(filename)


def check_toppers():
    wb = openpyxl.load_workbook('Rtu_Results_all.xlsx')
    ws = wb.get_sheet_by_name('Sheet')
    l = 68  # len(ws.rows)
    names = [ws['A' + str(i)].value for i in range(2, l + 1)]
    ch = 66
    for item in headings[1:-1]:
        marks = [[int(ws[chr(ch) + str(i)].value), i] for i in range(2, l + 1)]
        print 'Topper of ', item, 'is', names[max(marks, key=lambda x: x[0])[1] - 2]
        if 'Total' in item:
            print sorted()
        ch += 1


# save_to_excel('Rtu_Results_all.xlsx', pickle.load(open('All_Students.rtu', 'rb')))
if __name__ == '__main__':
    check_toppers()
    # save_to_excel('All_Students_with_deploma.xlsx', pickle.load(open('final_result.rtu', 'rb')))
